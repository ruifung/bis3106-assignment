from openpyxl import load_workbook
from itertools import islice
from functools import reduce
import pandas as pd

#Load the excel file
wb = load_workbook('./rainfall_with_state.xlsx', data_only=True, keep_vba=False, keep_links=False)
sheets = wb.worksheets

#Find the station detials (name, location, elevation)
def find_station_details(sheet):
    row = 1
    print(f'Loading station details from sheet: {sheet.title}')
    while row < sheet.max_row:
        if sheet.cell(row, 1).value == 'Station':
            break
        row += 1
    station_name = sheet.cell(row, 2).value.strip()[1:].strip()
    station_lat = sheet.cell(row + 1, 2).value.strip()[1:].strip()
    station_lon = sheet.cell(row + 2, 2).value.strip()[1:].strip()
    station_elevation = sheet.cell(row + 3, 1).value.split(':')[1].strip()
    return {
        'name': station_name,
        'latitude': station_lat,
        'longitude': station_lon,
        'elevation': station_elevation
    }

#Find the header row of the sheet
def findHeaderRow(sheet):
    row = 1
    while row < sheet.max_row:
        if sheet.cell(row, 1).value == 'Stnno':
            return row
        row += 1
        if row > 30:
            break
    return None

#Find the final data row of the sheet
def findFinalDataRow(sheet, headerRow):
    row = sheet.max_row
    stnno = sheet.cell(headerRow + 1, 1).value
    while row > 0:
        print(f"Trying cell({row},1) in sheet: {sheet.title}")
        cell = sheet.cell(row, 1)
        if cell and cell.value == stnno:
            print(f"Found row {row} as last row for sheet: {sheet.title}.")
            return row
        row -= 1
    return None

#Read the sheet into a dataframe
def toDataFrame(sheet, headerRow, finalDataRow):
    values = sheet.values
    data = islice(values, headerRow - 1, finalDataRow - 1)
    cols = next(data)
    data = list(data)
    df =  pd.DataFrame(data, columns=cols)
    station_details = find_station_details(sheet)
    df['station_name'] = station_details['name']
    df['station_latitude'] = station_details['latitude']
    df['station_longitude'] = station_details['longitude']
    df['station_elevation'] = station_details['elevation']
    return df

#Locate all header rows and final data rows
headerRows = [findHeaderRow(sheet) for sheet in sheets]
finalRows = [findFinalDataRow(sheet, headerRow) for sheet, headerRow in zip(sheets, headerRows)]

#Read all sheets into dataframes
dataframes = [toDataFrame(sheet, headerRow, finalRow) for sheet, headerRow, finalRow in zip(sheets, headerRows, finalRows)]

#Find all column names
all_column_names = set([column for dataframe in dataframes for column in list(dataframe.columns.values)])

#Append all sheets to form a single dataframe.
merged_df = reduce((lambda df1, df2: pd.concat([df1, df2], axis=0, ignore_index=True)), dataframes)

#Write output to CSV file.
merged_df.to_csv('rainfall_by_state.csv', index=False)